import streamlit as st
import pdfplumber
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import time
import re
st.title("Hämta ut info från rithuvud och granska")

st.markdown("""
Ladda upp ritningar och exportera info i rithuvud. Jämför ritningsnummer med filnamn, och granskar egna värden. Fungerar bara om filer är plottade rätt så rithuvud inte är förskjutet.  
v.2.8
""")

MM_TO_PT = 2.83465

BOXES_K2K3_MM = {
    "STATUS": (20, 110, 109, 122),
    "HANDLING": (20, 110, 100, 112),
    "DATUM": (89, 110, 94, 100),
    "ÄNDRING": (10, 90, 94, 100),
    "PROJEKT": (10, 112, 73, 93),
    "KONTAKTPERSON": (60, 110, 47, 52),
    "SKAPAD AV": (10, 60, 47, 52),
    "GODKÄND AV": (60, 110, 40, 45),
    "UPPDRAGSNUMMER": (10, 60, 40, 45),
    "RITNINGSKATEGORI": (28.6, 110, 33, 40),
    "INNEHÅLL": (57, 110, 19, 33),
    "FORMAT": (10, 30, 26, 31),
    "SKALA": (16, 30, 19, 26.5),
    "NUMMER": (39, 110, 10, 19),
    "BET": (14.7, 30, 10, 19)
}

BOXES_K1_MM = {
    "STATUS": (30, 120, 121, 131),
    "HANDLING": (30, 120, 110.5, 121),
    "DATUM": (100, 120, 104, 109),
    "ÄNDRING": (30, 100, 104, 109),
    "PROJEKT": (20, 120, 84, 102),
    "KONTAKTPERSON": (70, 120, 57, 62),
    "SKAPAD AV": (20, 70, 57, 62),
    "GODKÄND AV": (70, 120, 50, 55),
    "UPPDRAGSNUMMER": (20, 70, 50, 55),
    "RITNINGSKATEGORI": (38.6, 120, 43, 50),
    "INNEHÅLL": (67, 120, 29, 43),
    "FORMAT": (28.5, 41, 35, 42),
    "SKALA": (27.5, 41, 28, 36.5),
    "NUMMER": (49, 120, 18, 29.5),
    "BET": (28.5, 41, 18, 29.6)
}

BOXES_K12_MM = {
    "STATUS": (29.8, 119.8, 123, 133),
    "HANDLING": (29.8, 119.8, 112, 123),
    "DATUM": (93.5, 119.8, 106, 111),
    "ÄNDRING": (29.8, 93.5, 106, 111),
    "PROJEKT": (29, 119.8, 86, 106),
    "KONTAKTPERSON": (69.8, 119.8, 59, 64),
    "SKAPAD AV": (19.8, 69.8, 59, 64),
    "GODKÄND AV": (69.8, 119.8, 52, 57),
    "UPPDRAGSNUMMER": (19.8, 69.8, 52, 57),
    "RITNINGSKATEGORI": (38.4, 119.8, 45, 52),
    "INNEHÅLL": (66.8, 119.8, 31, 45),
    "FORMAT": (28.3, 40.8, 37, 44),
    "SKALA": (29, 40.8, 30, 37.5),
    "NUMMER": (48.8, 119.8, 20, 31.5),
    "BET": (28.3, 40.8, 20, 31.6)
}

coordinate_option = st.selectbox("Välj filstorlek för ritning", ["Helplan", "A1", "A1-5271"])
BOXES_MM = BOXES_K2K3_MM if coordinate_option == "Helplan" else BOXES_K1_MM if coordinate_option == "A1" else BOXES_K12_MM

uploaded_files = st.file_uploader("Ladda upp PDF", type="pdf", accept_multiple_files=True)
status_placeholder = st.empty()

# Input fields for comparison
st.subheader("Jämför med värden (valfritt)")
comparison_inputs = {}
for field in BOXES_MM.keys():
    comparison_inputs[field] = st.text_input(f"Förväntat värde för '{field}'", "")

def mm_box_to_pdf_bbox(page_width, page_height, x1_mm, x2_mm, y1_mm, y2_mm):
    x1_pt = page_width - x2_mm * MM_TO_PT
    x2_pt = page_width - x1_mm * MM_TO_PT
    y1_pt = page_height - y2_mm * MM_TO_PT
    y2_pt = page_height - y1_mm * MM_TO_PT
    return (x1_pt, y1_pt, x2_pt, y2_pt)

def extract_boxes(pdf_file, filename):
    extracted_rows = []
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            page_width = page.width
            page_height = page.height
            row = {"File": filename}

            # Extract predefined boxes
            for field, (x1_mm, x2_mm, y1_mm, y2_mm) in BOXES_MM.items():
                bbox = mm_box_to_pdf_bbox(page_width, page_height, x1_mm, x2_mm, y1_mm, y2_mm)
                cropped = page.within_bbox(bbox)
                text = cropped.extract_text()
                row[field] = text.strip() if text else ""

            # Scan bottom 10% for scale pattern
            bottom_bbox = (0, page_height * 0.9, page_width, page_height)
            bottom_crop = page.within_bbox(bottom_bbox)
            bottom_text = bottom_crop.extract_text() if bottom_crop else ""
            scale_matches = re.findall(r"1\s*:\s*\d+", bottom_text or "")
            row["Skalstock och skala"] = ", ".join(scale_matches) if scale_matches else ""

            extracted_rows.append(row)
    return extracted_rows

if st.button("Starta") and uploaded_files:
    status_placeholder.info("Körning pågår...")

    all_data = []
    progress_bar = st.progress(0)
    total_files = len(uploaded_files)

    for i, file in enumerate(uploaded_files):
        all_data.extend(extract_boxes(file, file.name))
        progress_bar.progress((i + 1) / total_files)

    df = pd.DataFrame(all_data)

    wb = Workbook()
    ws = wb.active
    ws.title = "Metadata"
    ws.append(df.columns.tolist())

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    for index, row in df.iterrows():
        excel_row = [row[col] for col in df.columns]
        ws.append(excel_row)

        for col in df.columns:
            expected = comparison_inputs.get(col, "").strip().lower()
            actual = str(row[col]).strip().lower()
            if expected:
                cell = ws.cell(row=ws.max_row, column=df.columns.get_loc(col) + 1)
                cell.fill = green_fill if actual == expected else red_fill

        # Extra comparison: filename vs NUMMER
        file_val = str(row["File"]).strip().lower().replace(".pdf", "")
        nummer_val = str(row["NUMMER"]).strip().lower()
        if file_val != nummer_val:
            cell = ws.cell(row=ws.max_row, column=df.columns.get_loc("NUMMER") + 1)
            cell.fill = red_fill

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    status_placeholder.success("Export och jämförelse färdig!")
    st.download_button(
        label="Ladda ner sammanfattning",
        data=output,
        file_name="metadata_comparison.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )




